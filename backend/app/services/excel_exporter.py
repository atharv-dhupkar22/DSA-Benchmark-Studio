from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Export benchmark history to an Excel workbook.
    """

    HEADERS = [
        "ID",
        "Created At",
        "Algorithm",
        "Category",
        "Dataset Type",
        "Dataset Size",
        "Runs",
        "Average (ms)",
        "Median (ms)",
        "Minimum (ms)",
        "Maximum (ms)",
        "Std Dev (ms)",
        "Peak Memory (KB)",
        "Time Complexity",
        "Space Complexity",
        "Stable",
        "In Place",
        "Python Version",
        "Benchmark Engine",
    ]

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
    )

    CENTER = Alignment(
        horizontal="center",
        vertical="center",
    )

    def export(
        self,
        output_path: Path,
        data: list[dict[str, Any]],
    ) -> Path:
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Benchmark Results"

        # -------------------------
        # Header
        # -------------------------
        for column, header in enumerate(self.HEADERS, start=1):
            cell = worksheet.cell(
                row=1,
                column=column,
                value=header,
            )

            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER

        # -------------------------
        # Data
        # -------------------------
        row_number = 2

        for item in data:

            execution = item.get("execution", {})
            metrics = item.get("algorithm_metrics", {})
            metadata = item.get("metadata", {})

            values = [
                item.get("id", ""),
                item.get("created_at", ""),
                item.get("algorithm", ""),
                item.get("category", ""),
                item.get("dataset_type", ""),
                item.get("dataset_size", ""),
                item.get("runs", ""),
                execution.get("average_ms", ""),
                execution.get("median_ms", ""),
                execution.get("minimum_ms", ""),
                execution.get("maximum_ms", ""),
                execution.get("standard_deviation_ms", ""),
                item.get("peak_memory_kb", ""),
                metrics.get("time_complexity", ""),
                metrics.get("space_complexity", ""),
                metrics.get("stable", ""),
                metrics.get("in_place", ""),
                metadata.get("python_version", ""),
                metadata.get("benchmark_engine", ""),
            ]

            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

                cell.alignment = self.CENTER

            row_number += 1

        # -------------------------
        # Auto-size columns
        # -------------------------
        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)

                if len(value) > max_length:
                    max_length = len(value)

            worksheet.column_dimensions[
                column_letter
            ].width = min(max_length + 4, 40)

        # -------------------------
        # Freeze Header
        # -------------------------
        worksheet.freeze_panes = "A2"

        workbook.save(output_path)

        return output_path